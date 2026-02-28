# pyright: reportUnknownMemberType=false, reportUnknownVariableType=false, reportUnknownArgumentType=false
"""Document text extraction for PDFs, Word, email, images, plain text.

All extraction functions are synchronous and called via asyncio.to_thread()
to avoid blocking the event loop during CPU-bound operations.
"""

from __future__ import annotations

import asyncio
import email as email_lib
from collections.abc import Callable
from pathlib import Path

from app.core.logging import get_logger
from app.knowledge.exceptions import UnsupportedDocumentTypeError

logger = get_logger(__name__)


async def extract_text(file_path: str, source_type: str) -> tuple[str, bool]:
    """Extract text from a document file.

    Routes to the appropriate extractor based on source_type.
    All extractors run in a thread pool to keep the event loop responsive.
    For PDFs, detects scanned pages and falls back to OCR if needed.

    Args:
        file_path: Absolute path to the file on disk.
        source_type: One of pdf, docx, email, image, text.

    Returns:
        Tuple of (extracted text, whether OCR was applied).

    Raises:
        UnsupportedDocumentTypeError: If source_type is not recognized.
    """
    logger.info("knowledge.extraction.started", source_type=source_type, file_path=file_path)

    # PDF is handled separately because it returns a tuple with OCR info
    if source_type == "pdf":
        text, ocr_applied = await asyncio.to_thread(_extract_pdf_sync, file_path)
        logger.info(
            "knowledge.extraction.completed",
            source_type=source_type,
            char_count=len(text),
            ocr_applied=ocr_applied,
        )
        return text, ocr_applied

    extractors: dict[str, Callable[[str], str]] = {
        "docx": _extract_docx_sync,
        "email": _extract_email_sync,
        "image": _extract_image_sync,
        "text": _extract_text_sync,
        "xlsx": _extract_excel_sync,
        "csv": _extract_csv_sync,
    }

    extractor = extractors.get(source_type)
    if extractor is None:
        raise UnsupportedDocumentTypeError(f"Unsupported document type: {source_type}")

    text = await asyncio.to_thread(extractor, file_path)

    logger.info(
        "knowledge.extraction.completed",
        source_type=source_type,
        char_count=len(text),
    )
    return text, False


def _extract_pdf_sync(file_path: str) -> tuple[str, bool]:
    """Extract text from a PDF file using PyMuPDF, with OCR fallback.

    If PyMuPDF extracts little or no text (< 50 chars total), assumes the PDF
    contains scanned images and falls back to pytesseract OCR.

    Args:
        file_path: Path to the PDF file.

    Returns:
        Tuple of (concatenated text from all pages, whether OCR was applied).
    """
    import fitz

    doc = fitz.open(file_path)
    pages: list[str] = []
    for page in doc:
        raw = page.get_text()
        page_text = str(raw) if raw else ""
        if page_text.strip():
            pages.append(page_text)

    total_text = "\n\n".join(pages)

    # If sufficient text was extracted, return without OCR
    if len(total_text.strip()) >= 50:
        doc.close()
        return total_text, False

    # Scanned PDF detected — fall back to OCR
    page_count = len(doc)
    logger.info(
        "knowledge.extraction.ocr_fallback",
        file_path=file_path,
        page_count=page_count,
        extracted_chars=len(total_text.strip()),
    )

    try:
        import pytesseract
        from PIL import Image

        ocr_pages: list[str] = []
        # Limit to first 50 pages to avoid memory issues
        max_ocr_pages = min(page_count, 50)
        for i in range(max_ocr_pages):
            page = doc[i]
            pix = page.get_pixmap(dpi=300)
            image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            ocr_text = str(pytesseract.image_to_string(image, lang="eng"))
            if ocr_text.strip():
                ocr_pages.append(ocr_text)

        doc.close()
        result = "\n\n".join(ocr_pages)
        logger.info(
            "knowledge.extraction.ocr_completed",
            char_count=len(result),
            pages_ocrd=max_ocr_pages,
        )
        return result, True

    except Exception as e:
        doc.close()
        logger.warning(
            "knowledge.extraction.ocr_failed",
            file_path=file_path,
            error=str(e),
            error_type=type(e).__name__,
            exc_info=True,
        )
        # Return whatever text we got (possibly empty)
        return total_text, False


def _extract_docx_sync(file_path: str) -> str:
    """Extract text from a Word document using python-docx.

    Args:
        file_path: Path to the .docx file.

    Returns:
        Concatenated paragraph text.
    """
    from docx import Document as DocxDocument

    doc = DocxDocument(file_path)
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_email_sync(file_path: str) -> str:
    """Extract text from an email file (.eml).

    Args:
        file_path: Path to the email file.

    Returns:
        Email headers and body text.
    """
    with Path(file_path).open() as f:
        msg = email_lib.message_from_file(f)

    parts: list[str] = []

    # Extract headers
    for header in ("From", "To", "Date", "Subject"):
        value = msg.get(header)
        if value:
            parts.append(f"{header}: {value}")

    # Extract body
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                if isinstance(payload, bytes):
                    parts.append(payload.decode("utf-8", errors="replace"))
    else:
        payload = msg.get_payload(decode=True)
        if isinstance(payload, bytes):
            parts.append(payload.decode("utf-8", errors="replace"))

    return "\n\n".join(parts)


def _extract_image_sync(file_path: str) -> str:
    """Extract text from an image using OCR (Tesseract).

    Args:
        file_path: Path to the image file.

    Returns:
        OCR-extracted text.
    """
    import pytesseract
    from PIL import Image

    image = Image.open(file_path)
    raw = pytesseract.image_to_string(image, lang="eng")
    return str(raw)


def _extract_text_sync(file_path: str) -> str:
    """Read a plain text file.

    Args:
        file_path: Path to the text file.

    Returns:
        File content as string.
    """
    return Path(file_path).read_text(encoding="utf-8")


def _extract_excel_sync(file_path: str) -> str:
    """Extract text from an Excel (.xlsx) file using openpyxl.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Tab-separated text from all sheets.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows():
            cells: list[str] = []
            for cell in row:
                cells.append(str(cell.value) if cell.value is not None else "")
            rows.append("\t".join(cells))
        sheet_text = f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows)
        sheets.append(sheet_text)
    wb.close()
    return "\n\n".join(sheets)


def _extract_csv_sync(file_path: str) -> str:
    """Extract text from a CSV file.

    Args:
        file_path: Path to the CSV file.

    Returns:
        Tab-separated text from all rows.
    """
    import csv

    try:
        with Path(file_path).open(encoding="utf-8") as f:
            reader = csv.reader(f)
            rows: list[str] = []
            for row in reader:
                rows.append("\t".join(row))
            return "\n".join(rows)
    except csv.Error:
        logger.warning("knowledge.extraction.csv_fallback", file_path=file_path)
        return Path(file_path).read_text(encoding="utf-8")
